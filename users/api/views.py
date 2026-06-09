from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from django.db import models
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter

from ..models import UserProfile, ClassInfo
from .serializers import (
    LoginSerializer,
    UserSerializer,
    UserCreateSerializer,
    UserUpdateSerializer,
    ChangePasswordSerializer,
    ClassInfoSerializer,
)
from ..permissions import IsAdmin, IsTeacher


class LoginAPIView(TokenObtainPairView):
    """自定义登录视图 - 扩展返回用户信息"""
    serializer_class = LoginSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        user = serializer.validated_data['user']
        
        # 生成 JWT Token
        refresh = RefreshToken.for_user(user)
        
        # 获取用户信息
        user_serializer = UserSerializer(user)
        
        return Response({
            'code': 200,
            'message': '登录成功',
            'data': {
                'access': str(refresh.access_token),
                'refresh': str(refresh),
                'user': user_serializer.data
            }
        }, status=status.HTTP_200_OK)


class LogoutAPIView(APIView):
    """注销视图 - 将 refresh token 加入黑名单"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get('refresh')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
            
            return Response({
                'code': 200,
                'message': '注销成功'
            })
        except Exception as e:
            return Response({
                'code': 400,
                'message': f'注销失败: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)


class CurrentUserAPIView(APIView):
    """获取当前登录用户信息"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user)
        return Response({
            'code': 200,
            'message': 'success',
            'data': serializer.data
        })

    def put(self, request):
        """更新当前用户信息"""
        serializer = UserUpdateSerializer(
            request.user,
            data=request.data,
            partial=True,
            context={'request': request}
        )

        if serializer.is_valid():
            serializer.save()
            return Response({
                'code': 200,
                'message': '更新成功',
                'data': UserSerializer(request.user).data
            })

        return Response({
            'code': 400,
            'message': '数据验证失败',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        """上传头像"""
        if 'avatar' not in request.FILES:
            return Response({
                'code': 400,
                'message': '请选择要上传的头像文件'
            }, status=status.HTTP_400_BAD_REQUEST)

        avatar_file = request.FILES['avatar']

        if avatar_file.size > 5 * 1024 * 1024:
            return Response({
                'code': 400,
                'message': '头像文件大小不能超过 5MB'
            }, status=status.HTTP_400_BAD_REQUEST)

        allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp']
        if avatar_file.content_type not in allowed_types:
            return Response({
                'code': 400,
                'message': '只支持 JPG、PNG、GIF、WebP 格式的图片'
            }, status=status.HTTP_400_BAD_REQUEST)

        request.user.avatar = avatar_file
        request.user.save()

        return Response({
            'code': 200,
            'message': '头像上传成功',
            'data': {
                'avatar': request.user.avatar.url if request.user.avatar else None
            }
        })


class ChangePasswordAPIView(APIView):
    """修改密码"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ChangePasswordSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            user = request.user
            new_password = serializer.validated_data['new_password']
            user.set_password(new_password)
            user.must_change_password = False
            user.save()
            
            return Response({
                'code': 200,
                'message': '密码修改成功，请重新登录'
            })
        
        return Response({
            'code': 400,
            'message': '数据验证失败',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


class MyClassesAPIView(APIView):
    """获取当前教师管理的班级列表"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user

        if getattr(user, 'role', '') not in ['admin', 'teacher']:
            return Response({'code': 403, 'message': '无权限'}, status=403)

        classes = ClassInfo.objects.filter(
            students__teacher=user
        ).annotate(
            student_count=models.Count('students')
        ).distinct().values(
            'id', 'name', 'grade', 'major', 'class_number', 'student_count'
        )

        return Response({
            'code': 200,
            'data': list(classes)
        })


class UserViewSet(viewsets.ModelViewSet):
    """
    用户管理 ViewSet - 完整的 CRUD + 批量操作
    
    list: 返回用户列表（支持搜索、筛选、排序）
    retrieve: 返回用户详情
    create: 创建新用户（仅管理员）
    update: 更新用户信息（管理员或本人）
    partial_update: 部分更新
    destroy: 删除用户（仅管理员）
    
    自定义 action:
    - bulk_delete: 批量删除
    - reset_password: 重置密码
    - my_students: 获取我的学生列表（教师专用）
    """
    
    queryset = UserProfile.objects.all()
    serializer_class = UserSerializer
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['role', 'is_active', 'major', 'semester', 'class_info']
    search_fields = ['username', 'email', 'student_id', 'phone', 'first_name', 'last_name']
    ordering_fields = ['date_joined', 'last_login', 'username', 'id']
    ordering = ['-date_joined']

    def get_permissions(self):
        """根据 action 动态设置权限"""
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated]
        elif self.action == 'create':
            permission_classes = [IsAdmin]
        elif self.action in ['update', 'partial_update']:
            permission_classes = [IsAuthenticated]
        elif self.action == 'destroy':
            permission_classes = [IsAdmin]
        elif self.action in ['bulk_delete', 'reset_password']:
            permission_classes = [IsAdmin]
        elif self.action == 'my_students':
            permission_classes = [IsTeacher]
        else:
            permission_classes = [IsAuthenticated]
        
        return [permission() for permission in permission_classes]

    def get_queryset(self):
        """根据用户角色过滤数据"""
        user = self.request.user
        
        if not user.is_authenticated:
            return UserProfile.objects.none()
        
        if user.role == 'admin':
            queryset = UserProfile.objects.all()
        elif user.role == 'teacher':
            queryset = UserProfile.objects.filter(
                models.Q(pk=user.pk) | models.Q(teacher=user)
            )
        else:
            queryset = UserProfile.objects.filter(pk=user.pk)
        
        return queryset

    def get_serializer_class(self):
        """根据 action 选择序列化器"""
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        elif self.action == 'change_password':
            return ChangePasswordSerializer
        return UserSerializer

    def list(self, request):
        """返回用户列表（包装响应格式）"""
        queryset = self.filter_queryset(self.get_queryset())
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            paginator = self.paginator
            return Response({
                'code': 200,
                'message': 'success',
                'data': {
                    'results': serializer.data,
                    'count': paginator.page.paginator.count,
                    'next': paginator.get_next_link(),
                    'previous': paginator.get_previous_link(),
                    'page': paginator.page.number,
                    'total_pages': paginator.page.paginator.num_pages
                }
            })
        
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'code': 200,
            'message': 'success',
            'data': {
                'results': serializer.data,
                'count': queryset.count()
            }
        })

    def create(self, request, *args, **kwargs):
        """创建用户（仅管理员）"""
        serializer = self.get_serializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.save()
            
            return Response({
                'code': 201,
                'message': '用户创建成功',
                'data': UserSerializer(user).data
            }, status=status.HTTP_201_CREATED)
        
        return Response({
            'code': 400,
            'message': '创建失败',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request):
        """更新用户信息（支持管理员或本人）"""
        instance = self.get_object()
        
        # 权限检查：非管理员只能修改自己的信息
        if request.user.role != 'admin' and instance.pk != request.user.pk:
            return Response({
                'code': 403,
                'message': '您没有权限修改其他用户的信息'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = self.get_serializer(instance, data=request.data, partial=True, context={'request': request})
        
        if serializer.is_valid():
            user = serializer.save()
            
            return Response({
                'code': 200,
                'message': '更新成功',
                'data': UserSerializer(user).data
            })
        
        return Response({
            'code': 400,
            'message': '数据验证失败',
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """批量删除用户"""
        ids = request.data.get('ids', [])
        
        if not ids:
            return Response({
                'code': 400,
                'message': '请提供要删除的用户 ID 列表'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        deleted_count, _ = UserProfile.objects.filter(id__in=ids).delete()
        
        return Response({
            'code': 200,
            'message': f'成功删除 {deleted_count} 个用户'
        })

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """重置用户密码"""
        user = self.get_object()
        new_password = request.data.get('new_password')
        
        if not new_password:
            return Response({
                'code': 400,
                'message': '请提供新密码'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if len(new_password) < 6:
            return Response({
                'code': 400,
                'message': '密码长度不能少于6位'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user.set_password(new_password)
        user.save()
        
        return Response({
            'code': 200,
            'message': f'用户 {user.username} 的密码已重置成功'
        })

    @action(detail=False, methods=['get'])
    def my_students(self, request):
        """获取当前教师的学生列表"""
        teacher = request.user

        if teacher.role not in ['admin', 'teacher']:
            return Response({
                'code': 403,
                'message': '只有教师才能查看学生列表'
            }, status=status.HTTP_403_FORBIDDEN)

        students = UserProfile.objects.filter(teacher=teacher)
        
        # 应用搜索
        search = request.query_params.get('search')
        if search:
            from django.db.models import Q
            students = students.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(student_id__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
        
        # 应用排序
        ordering = request.query_params.get('ordering', '-date_joined')
        students = students.order_by(ordering)
        
        # 分页
        page = self.paginate_queryset(students)
        if page is not None:
            serializer = UserSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = UserSerializer(students, many=True)
        return Response({
            'code': 200,
            'data': serializer.data,
            'total': students.count()
        })

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """获取用户统计数据"""
        total_users = UserProfile.objects.count()
        admin_count = UserProfile.objects.filter(role='admin').count()
        teacher_count = UserProfile.objects.filter(role='teacher').count()
        student_count = UserProfile.objects.filter(role='student').count()

        recent_users = UserProfile.objects.order_by('-date_joined')[:5]

        # 班级维度统计
        class_stats = []
        for cls in ClassInfo.objects.all():
            students = UserProfile.objects.filter(class_info=cls, role='student')
            class_stats.append({
                'class_id': cls.id,
                'class_name': cls.name,
                'student_count': students.count(),
            })

        return Response({
            'code': 200,
            'data': {
                'total': total_users,
                'by_role': {
                    'admin': admin_count,
                    'teacher': teacher_count,
                    'student': student_count
                },
                'by_class': class_stats,
                'recent_users': UserSerializer(recent_users, many=True).data
            }
        })

    @action(detail=True, methods=['put'])
    def change_status(self, request, pk=None):
        """启用/禁用用户账号"""
        user = self.get_object()
        is_active = request.data.get('is_active')
        
        if is_active is None:
            return Response({
                'code': 400,
                'message': '请提供 is_active 参数 (true/false)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user.is_active = is_active
        user.save()
        
        status_text = '启用' if is_active else '禁用'
        
        return Response({
            'code': 200,
            'message': f'用户 {user.username} 已{status_text}',
            'data': UserSerializer(user).data
        })

    @action(detail=False, methods=['post'])
    def preview_import(self, request):
        """预览导入：解析 Excel，返回数据供用户编辑确认（不写库）"""
        if 'file' not in request.FILES:
            return Response({
                'code': 400,
                'message': '请上传 Excel 文件'
            }, status=status.HTTP_400_BAD_REQUEST)

        file = request.FILES['file']
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({
                'code': 400,
                'message': '仅支持 .xlsx/.xls 格式'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
        except Exception as e:
            return Response({
                'code': 400,
                'message': f'文件解析失败: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

        rows = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            role_raw = str(row[0] or '').strip().lower()
            identifier = str(row[1] or '').strip()
            name = str(row[2] or '').strip()
            class_name = str(row[3] or '').strip() or ''
            phone = str(row[4] or '').strip() or ''
            semester = str(row[5] or '').strip() or ''
            teacher_name = str(row[6] or '').strip() or ''

            # 基础校验
            if not role_raw or not identifier or not name:
                errors.append({'row': row_idx, 'error': '角色、标识符和姓名为必填项'})
                continue

            if role_raw not in ('student', 'teacher'):
                errors.append({'row': row_idx, 'error': f'无效角色: {role_raw}'})
                continue

            rows.append({
                'row': row_idx,
                'role': role_raw,
                'identifier': identifier,
                'name': name,
                'class_name': class_name,
                'phone': phone,
                'semester': semester,
                'teacher_name': teacher_name,
            })

        # 提供已有班级和教师列表供前端下拉选择
        class_list = list(ClassInfo.objects.values('id', 'name'))
        teacher_list = list(UserProfile.objects.filter(
            role__in=['admin', 'teacher']
        ).values('id', 'real_name', 'username').order_by('real_name'))

        return Response({
            'code': 200,
            'message': f'解析完成，共 {len(rows)} 条数据',
            'data': {
                'rows': rows,
                'errors': errors,
                'class_options': class_list,
                'teacher_options': teacher_list,
                'total': len(rows),
            }
        })

    @action(detail=False, methods=['post'])
    def confirm_import(self, request):
        """确认导入：接收前端编辑后的数据，写入数据库"""
        try:
            rows_data = request.data.get('rows', [])
            if not rows_data:
                return Response({
                    'code': 400,
                    'message': '没有要导入的数据'
                }, status=status.HTTP_400_BAD_REQUEST)

            results = []
            errors = []
            created_classes = {}

            for item in rows_data:
                row_idx = item.get('row', '?')
                try:
                    role_raw = item.get('role', '')
                    identifier = item.get('identifier', '')
                    name = item.get('name', '')
                    class_name = item.get('class_name', '') or ''
                    phone = item.get('phone', '') or ''
                    semester = item.get('semester', '') or ''
                    teacher_name = item.get('teacher_name', '') or ''

                    if role_raw not in ('student', 'teacher'):
                        errors.append({'row': row_idx, 'error': f'无效角色: {role_raw}'})
                        continue

                    # 班级自动创建（学生）
                    class_obj = None
                    if class_name and role_raw == 'student':
                        if class_name in created_classes:
                            class_obj = created_classes[class_name]
                        else:
                            class_obj, _ = ClassInfo.objects.get_or_create(
                                name=class_name,
                                defaults={'grade': '', 'major': '', 'class_number': ''}
                            )
                            created_classes[class_name] = class_obj

                    # 指导教师匹配（按姓名或用户名）
                    teacher_obj = None
                    if teacher_name:
                        try:
                            teacher_obj = UserProfile.objects.filter(
                                models.Q(real_name=teacher_name) | models.Q(username=teacher_name),
                                role__in=['admin', 'teacher']
                            ).first()
                        except Exception:
                            pass

                    # 唯一性检查
                    if role_raw == 'student':
                        if UserProfile.objects.filter(student_id=identifier).exists():
                            errors.append({'row': row_idx, 'error': f'学号 {identifier} 已存在'})
                            continue
                    else:
                        if UserProfile.objects.filter(employee_id=identifier).exists():
                            errors.append({'row': row_idx, 'error': f'工号 {identifier} 已存在'})
                            continue

                    user_data = {
                        'username': identifier,
                        'real_name': name,
                        'role': role_raw,
                        'phone': phone,
                        'semester': semester,
                        'class_info': class_obj,
                        'teacher': teacher_obj,
                    }
                    if role_raw == 'student':
                        user_data['student_id'] = identifier
                    else:
                        user_data['employee_id'] = identifier

                    user = UserProfile(**user_data)
                    default_pwd = phone[-6:] if len(phone) >= 6 else (identifier[-6:] if len(identifier) >= 6 else '123456')
                    user.set_password(default_pwd)
                    user.must_change_password = True
                    user.save()

                    results.append({
                        'row': row_idx,
                        'role': role_raw,
                        'identifier': identifier,
                        'name': name,
                        'class_name': class_name or '-',
                        'phone': phone or '-',
                        'default_password': default_pwd,
                        'status': 'success',
                    })

                except Exception as e:
                    errors.append({'row': row_idx, 'error': str(e)})

            return Response({
                'code': 200,
                'message': f'导入完成：成功 {len(results)} 条，失败 {len(errors)} 条',
                'data': {
                    'results': results,
                    'errors': errors,
                    'total_success': len(results),
                    'total_errors': len(errors),
                }
            })
        except Exception as e:
            return Response({
                'code': 500,
                'message': f'导入异常: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def download_template(self):
        """下载导入模板"""
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from django.http import HttpResponse
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '用户导入模板'

        headers = ['角色', '学号/工号', '姓名', '班级名称', '手机号', '学期', '指导教师']
        ws.append(headers)

        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 示例数据行
        ws.append(['student', '202301001', '张三', '计算机科学2301班', '13800138000', '2024-2025-1', '王教授'])
        ws.append(['teacher', 'T10001', '李老师', '', '13900139000', '', ''])

        # 调整列宽
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="user_import_template.xlsx"'
        return response


class ClassInfoViewSet(viewsets.ModelViewSet):
    """班级管理 ViewSet"""
    queryset = ClassInfo.objects.all().order_by('grade', 'major', 'class_number')
    serializer_class = ClassInfoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['grade', 'major']
    search_fields = ['name', 'major']
    ordering_fields = ['grade', 'name']

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAdmin]
        return [p() for p in permission_classes]

    def list(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True)
        return Response({
            'code': 200,
            'message': 'success',
            'data': {'results': serializer.data, 'count': queryset.count()}
        })


# ============================================================
# 教师工作台 API
# ============================================================

from django.utils import timezone
from django.db.models import Avg, Count, Max, Q
from datetime import timedelta


class TeacherWorkspaceView(APIView):
    """教师工作台 - 班级学生数据概览"""
    permission_classes = [IsAuthenticated]

    def _get_date_range(self, period):
        now = timezone.now()
        if period == '7d':
            return now - timedelta(days=7), now
        elif period == '90d':
            return now - timedelta(days=90), now
        else:
            return now - timedelta(days=30), now

    def _get_managed_students(self, user, class_id=None):
        qs = UserProfile.objects.filter(teacher=user)
        if class_id:
            qs = qs.filter(class_info_id=class_id)
        return qs

    def get(self, request):
        from practice.models import PracticeRecord

        user = request.user
        role = getattr(user, 'role', '')

        if role not in ['admin', 'teacher']:
            return Response({'code': 403, 'message': '无权限'}, status=403)

        target_teacher = user

        class_id = request.query_params.get('class_id')
        period = request.query_params.get('period', '30d')

        students_qs = self._get_managed_students(target_teacher, class_id)
        student_ids = list(students_qs.values_list('pk', flat=True))

        if not student_ids:
            return Response({
                'code': 200,
                'data': {
                    'stats': {
                        'total_students': 0,
                        'class_avg_score': 0,
                        'material_completion_rate': 0,
                        'active_rate': 0,
                        'total_practices': 0,
                        'attention_count': 0,
                    },
                    'students': [],
                    'charts': {'score_distribution': {}, 'activity_trend': []},
                    'alerts': [],
                }
            })

        start_date, end_date = self._get_date_range(period)

        # B区：统计卡片
        total_students = len(student_ids)

        practice_qs = PracticeRecord.objects.filter(
            user_id__in=student_ids,
            created_at__gte=start_date,
            created_at__lte=end_date,
            is_completed=True
        )
        score_agg = practice_qs.aggregate(avg=Avg('overall_score'))
        class_avg_score = round(score_agg['avg'] or 0, 1)
        total_practices = practice_qs.count()

        seven_days_ago = timezone.now() - timedelta(days=7)
        active_count = UserProfile.objects.filter(
            pk__in=student_ids,
            last_login__gte=seven_days_ago
        ).count()
        active_rate = round(active_count / total_students * 100) if total_students > 0 else 0

        # C区：学生列表 + 聚合数据
        student_stats = {}
        records_by_student = practice_qs.values('user_id').annotate(
            count=Count('id'),
            avg_score=Avg('overall_score'),
        )
        for r in records_by_student:
            student_stats[r['user_id']] = {
                'practice_count': r['count'],
                'avg_score': round(r['avg_score'] or 0, 1),
            }

        class_avg_practices = total_practices / total_students if total_students > 0 else 0

        students_data = []
        alerts = []

        for s in students_qs.select_related('class_info'):
            stats = student_stats.get(s.pk, {'practice_count': 0, 'avg_score': 0})
            avg_score = stats['avg_score']
            practices = stats['practice_count']

            # 状态判定
            days_since_login = (timezone.now() - s.last_login).days if s.last_login else 999
            if avg_score >= 85 and practices >= class_avg_practices:
                status = 'excellent'
            elif avg_score >= 60 and days_since_login <= 7:
                status = 'normal'
            else:
                status = 'attention'

            student_entry = {
                'id': s.pk,
                'real_name': s.real_name or s.username,
                'student_id': s.student_id or '',
                'class_name': s.class_info.name if s.class_info else '',
                'completion_rate': 0,
                'practice_count': practices,
                'avg_score': avg_score,
                'last_login': s.last_login.isoformat() if s.last_login else None,
                'status': status,
            }
            students_data.append(student_entry)

            # E区：预警规则
            if status == 'attention':
                alert_reasons = []
                if avg_score < class_avg_score - 10 and class_avg_score > 0:
                    diff = round(class_avg_score - avg_score, 1)
                    alert_reasons.append({
                        'reason': 'score_low',
                        'detail': f'练习均分 {avg_score}，低于班级平均 {diff} 分',
                    })
                if days_since_login > 7:
                    alert_reasons.append({
                        'reason': 'inactive',
                        'detail': f'已连续 {days_since_login} 天未登录',
                    })
                if practices < class_avg_practices * 0.3 and class_avg_practices > 0:
                    alert_reasons.append({
                        'reason': 'low_practice',
                        'detail': f'练习次数偏少：{practices}次（班级平均 {round(class_avg_practices, 1)}次）',
                    })

                if alert_reasons:
                    alerts.append({
                        'student_id': s.pk,
                        'student_name': s.real_name or s.username,
                        **alert_reasons[0],
                        'avg_score': avg_score,
                        'last_login': s.last_login.isoformat() if s.last_login else None,
                    })

        attention_count = sum(1 for s in students_data if s['status'] == 'attention')

        # D区：图表数据
        distribution = {'excellent': 0, 'good': 0, 'average': 0, 'pass': 0, 'fail': 0}
        for s in students_data:
            sc = s['avg_score']
            if sc >= 90:
                distribution['excellent'] += 1
            elif sc >= 80:
                distribution['good'] += 1
            elif sc >= 70:
                distribution['average'] += 1
            elif sc >= 60:
                distribution['pass'] += 1
            else:
                distribution['fail'] += 1

        trend_start = timezone.now() - timedelta(days=30)
        daily_active = UserProfile.objects.filter(
            pk__in=student_ids,
            last_login__gte=trend_start
        ).extra(select={'day': 'date(last_login)'}).values('day').annotate(
            c=Count('id')
        ).order_by('day')

        daily_practices = PracticeRecord.objects.filter(
            user_id__in=student_ids,
            created_at__gte=trend_start
        ).extra(select={'day': 'date(created_at)'}).values('day').annotate(
            c=Count('id')
        ).order_by('day')

        activity_trend = []
        day_map = {}
        for d in daily_active:
            day_map[d['day'].isoformat()] = {'active_students': d['c'], 'practice_count': 0}
        for d in daily_practices:
            key = d['day'].isoformat()
            if key in day_map:
                day_map[key]['practice_count'] = d['c']
            else:
                day_map[key] = {'active_students': 0, 'practice_count': d['c']}
        for day_str in sorted(day_map.keys()):
            activity_trend.append({'date': day_str, **day_map[day_str]})

        return Response({
            'code': 200,
            'data': {
                'stats': {
                    'total_students': total_students,
                    'class_avg_score': class_avg_score,
                    'material_completion_rate': 65,
                    'active_rate': active_rate,
                    'total_practices': total_practices,
                    'attention_count': attention_count,
                },
                'students': students_data,
                'charts': {
                    'score_distribution': distribution,
                    'activity_trend': activity_trend,
                },
                'alerts': alerts,
            }
        })


class TeacherStudentDetailView(APIView):
    """教师工作台 - 学生详情"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        from practice.models import PracticeRecord

        user = request.user
        role = getattr(user, 'role', '')

        if role not in ['admin', 'teacher']:
            return Response({'code': 403, 'message': '无权限'}, status=403)

        try:
            student = UserProfile.objects.get(pk=pk)
        except UserProfile.DoesNotExist:
            return Response({'code': 404, 'message': '学生不存在'}, status=404)

        if role == 'teacher' and student.teacher != user:
            return Response({'code': 403, 'message': '无权查看此学生'}, status=403)

        period = request.query_params.get('period', '30d')
        now = timezone.now()
        if period == '7d':
            start_date = now - timedelta(days=7)
        elif period == '90d':
            start_date = now - timedelta(days=90)
        else:
            start_date = now - timedelta(days=30)

        basic = {
            'id': student.pk,
            'real_name': student.real_name or student.username,
            'student_id': student.student_id or '',
            'class_name': student.class_info.name if student.class_info else '',
            'avatar': student.avatar.url if student.avatar else None,
            'status': '',
            'date_joined': student.date_joined.isoformat() if student.date_joined else None,
        }

        all_records = PracticeRecord.objects.filter(user=student, is_completed=True)
        period_records = all_records.filter(
            created_at__gte=start_date, created_at__lte=now
        )

        practice_stats = all_records.aggregate(
            count=Count('id'), avg=Avg('overall_score'), max_score=Max('overall_score')
        )

        thirty_ago = now - timedelta(days=30)
        recent_active_dates = set(all_records.filter(
            created_at__gte=thirty_ago
        ).values_list('created_at__date', flat=True))

        stats = {
            'practice_count': practice_stats['count'] or 0,
            'avg_score': round(practice_stats['avg'] or 0, 1),
            'max_score': practice_stats['max_score'] or 0,
            'completion_rate': 0,
            'active_days': len(recent_active_dates),
        }

        avg_s = stats['avg_score']
        pc = stats['practice_count']
        days_login = (now - student.last_login).days if student.last_login else 999
        if avg_s >= 85:
            basic['status'] = 'excellent'
        elif avg_s >= 60 and days_login <= 7:
            basic['status'] = 'normal'
        else:
            basic['status'] = 'attention'

        trend_records = period_records.extra(
            select={'day': 'date(created_at)'}
        ).values('day').annotate(avg=Avg('overall_score')).order_by('day')
        score_trend = [
            {'date': t['day'].isoformat(), 'avg_score': round(t['avg'] or 0, 1)}
            for t in trend_records
        ]

        record_list = period_records.order_by('-created_at')[:50].select_related('scenario', 'topic')
        practice_records_data = [
            {
                'scenario_title': r.scenario.title if r.scenario else '未知场景',
                'topic_title': r.topic.title if r.topic else '',
                'overall_score': r.overall_score,
                'duration_seconds': r.duration_seconds,
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M'),
            }
            for r in record_list
        ]

        return Response({
            'code': 200,
            'data': {
                'basic': basic,
                'stats': stats,
                'score_trend': score_trend,
                'practice_records': practice_records_data,
            }
        })
